'''
****Data Structure****
data = {
    "india":{
        "Gujarat": ["Palanpur","Ahmedabad","Surat","Mahesana","Vadodara"],
        "Rajasthan": ["Jaipur","Udaipur","Ajmer","Jodhpur","Bikaner"],
        "Maharashtra": ["Mumbai","Pune","Nagpur","Nashik","Kolhapur"]
    },
    "Australia":{
    	"South Australia": ["Sydney"]
    }
}
'''
#implementation of the excel task. Line no : 29 - 96
import openpyxl
from openpyxl import Workbook, load_workbook
import os
data = {}
filename = "country_state_city.xlsx"

def insert_data_into_excel(data, filename):
    """"
    this function insert data into the excel file.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Country", "State", "City"])
    for country, states in data.items():
        if not states:  # add the country only
            sheet.append([country, "", ""])
        else:
            first_state = True
            for state, cities in states.items():
                if not cities:  # add the state only
                    if first_state:
                        sheet.append([country, state, ""])
                        first_state = False
                    else:
                        sheet.append(["", state, ""])
                else:
                    first_city = True
                    for city in cities:
                        if first_city:
                            if first_state:
                                sheet.append([country, state, city])
                                first_state = False
                            else:
                                sheet.append(["", state, city])
                            first_city = False
                        else:
                            sheet.append(["", "", city])
        sheet.append(["", "", ""])
    workbook.save(filename)
    print(f"Data successfully written to {filename}")

def retrieve_data_from_excel(filename):
    """"
        this function retrieve data from the excel file.
    """
    if not os.path.exists(filename):
        return {}
    workbook = load_workbook(filename)
    sheet = workbook.active
    data = {}
    current_country = None
    current_state = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        country, state, city = row
        if country:
            current_country = country
            data[current_country] = {}
        if state:
            current_state = state
            data[current_country][current_state] = []
        if city:
            data[current_country][current_state].append(city)
    return data

def continueOrNot():
    """"
        this function return true if user input y and return false if user input n.
        which means know the user want continue or not
    """
    while True:
        try:
            choice = str(input("Do you want to continue? (y/n): "))
            if choice == "y":
                return True
            elif choice == "n":
                return False
            else:
                print("Invalid input. Please enter 'y' or 'n'.")
        except ValueError:
            print("Invalid input. Please enter 'y' or 'n'.")

def continueOrNotWithMessage(message):
    """"
            this function return true if user input y and return false if user input n.
            which means know the user want continue or not
            we can additional parameter set message.
    """
    while True:
        try:
            choice = str(input(f"Do you want to {message}? (y/n): "))
            if choice == "y":
                return True
            elif choice == "n":
                return False
            else:
                print("Invalid input. Please enter 'y' or 'n'.")
        except ValueError:
            print("Invalid input. Please enter 'y' or 'n'.")

def validate_to_string(string):
    """"
        this function validating the string. string can not contain numbers.
    """
    if not string.isalpha():
        print("Please Enter string !!!")
        return False
    return True

def is_country(country):
    """"
        to_check the is it country or not. if country does not exists in data return False
    """
    if country not in data:
        print("Country not found. Please try again.")
        return False
    return True

def is_state(country, state):
    """"
    to check the state is exists or not corrensponding the country.
    """
    if state not in data[country]:
        print("State not found. Please try again.")
        return False
    return True

def is_city(country, state, city):
    """"
    to check city is exists or not corrensponding state and country
    """
    if city not in data[country][state]:
        print("City not found. Please try again.")
        return False
    return True

def add_country(country):
    """"
    this function convert country in lower and add them to data
    """
    if country.lower() in data:
        print('country already exits')
    else:
        data[country.lower()] = {}

def Display_state(country):
    """
    display the state corrensponding country
    """
    print("States: ", list(data[country].keys()))

def Display_city(country, state):
    """
    display the city corrensponding state and country
    """
    print("cities: ", list(data[country][state]))

def Display_country():
    """
    this function to display countries
    """
    print("Countries: ", list(data.keys()))

def is_country_empty():
    """
    this function checking is data has countries or not. if data has not any country return false.
    """
    if not bool(data):
        print("empty!!! country not found")
        return False
    return True

def is_city_empty(country, state):
    """
    this function checking is state has cities or not.
    """
    if not bool(data[country][state]):
        print("empty!!! city not found")
        return False
    return True

def is_state_empty(country):
    """
    this function checking is country has states or not.
    """
    if not bool(data[country]):
        print("empty!!! state not found")
        return False
    return True

def input_state():
    """
    this function get state from the user.
    """
    while True:
        state = str(input("Enter state name: "))
        val =  validate_to_string(state)
        if not val:
            continue
        return state.lower()

def input_country():
    """
    this fuction get country from the user.
    """
    while True:
        country = str(input("Enter Country name: "))
        val = validate_to_string(country)
        if not val:
            continue
        return country.lower()
    	 
def input_city():
    """
    this function get city from the user
    """
    while True:
        city = str(input("Enter City name: "))
        val = validate_to_string(city)
        if not val:
            continue
        return city.lower()

def continue_adding_state():
    """
    this function for continuesly adding state.
    """
    while True:
        Display_country()
        if not is_country_empty(): 
            break
        country = str(input("select country name: "))
        if not is_country(country.lower()):
            continue
        state = input_state()
        if state.lower() in data[country.lower()]:
            print("state is already exits!!!")
            continue
        data[country.lower()][state.lower()] = []
        print(data)
        corn= continueOrNot()
        if corn:
            continue
        else:
            break
#execution starts here-----------------
data = retrieve_data_from_excel(filename)       
while True:
    print("""
        Select operation:
           1. Add
           2. Update
           3. Delete
           4. Exit
           5. Display
    """)
    try:
        choice = int(input("Enter choice: "))
        if choice == 1:
            while True:
                try:
                    print("-----------adding country-------------")
                    country = input_country()
                    add_country(country) 
                    print(data)
                    print("country added......")
                    corn= continueOrNotWithMessage("enter another country")
                    if corn:
                        continue
                    else:
                        break
                except ValueError:
                    print("Invalid input. Please enter a string.")
                    continue
            while True:
                print("-----------adding state-------------")        
                Display_country()
                if not is_country_empty(): 
                    break
                country = str(input("select country name: "))
                if not is_country(country.lower()):
                    continue
                state = input_state()
                if state.lower() in data[country.lower()]:
                    print("state is already exits!!!")
                    corn = continueOrNotWithMessage("enter another state")
                    if corn:
                        continue
                    else:
                        break
                data[country.lower()][state.lower()] = []
                print(data)
                print("state added......")
                corn= continueOrNotWithMessage("enter another state")
                if corn:
                    continue
                else:
                    break
            while True: 
                print("-----------adding city-------------")
                Display_country()
                if not is_country_empty():
                    break
                country = str(input("select country name: "))
                if not is_country(country.lower()):
                    continue
                if not is_state_empty(country.lower()):
                    break
                Display_state(country.lower())
                state = str(input("select state name: "))
                if not is_state(country.lower(), state.lower()):
                    continue
                city = input_city()
                if city.lower() in data[country.lower()][state.lower()]:
                    print("city is already exits!!!")
                    corn = continueOrNotWithMessage("enter another city")
                    if corn:
                        continue
                    else:
                        break
                data[country.lower()][state.lower()].append(city.lower())
                print(data)
                print("city added......")
                corn= continueOrNotWithMessage("enter another city")
                if corn:
                    continue
                else:
                    break
        elif choice == 2:
            while True:
                print("""
                    ****Update*****
                      Select option:
                        1. Country
                        2. State
                        3. City
                        4. Exit
                """)
                try:
                    sub_choice = int(input("Enter choice: "))
                    if sub_choice == 1:
                        while True:
                            Display_country()
                            if not is_country_empty(): 
                                break
                            country = str(input("select country name: "))
                            if not is_country(country.lower()):
                                continue
                            new_country = input_country()
                            if new_country in data.keys():
                                print("Please Try another country name do not use same name or available country name")
                                break
                            data[new_country.lower()] = data.pop(country.lower())
                            print(data)
                            print("country updated......")
                            corn= continueOrNot()
                            if corn:
                                continue
                            else:
                                break
                    elif sub_choice == 2:
                        while True:
                            Display_country()
                            if not is_country_empty(): 
                                break
                            country = str(input("select country name: "))
                            if not is_country(country.lower()):
                                continue
                            if not is_state_empty(country.lower()):
                                break
                            Display_state(country.lower())
                            state = str(input("select state name: "))
                            if not is_state(country.lower(), state.lower()):
                                continue
                            new_state = input_state()
                            data[country.lower()][new_state.lower()] = data[country.lower()].pop(state.lower())
                            print(data)
                            print("state updated......")
                            corn= continueOrNot()
                            if corn:
                                continue
                            else:
                                break
                    elif sub_choice == 3:
                        while True:
                            Display_country()
                            if not is_country_empty(): 
                                break
                            country = str(input("select country name: "))
                            if not is_country(country.lower()):
                                continue
                            Display_state(country.lower())
                            if not is_state_empty(country.lower()):
                                break
                            state = str(input("select state name: "))
                            if not is_state(country.lower(), state.lower()):
                                continue
                            Display_city(country.lower(), state.lower())
                            if not is_city_empty(country.lower(), state.lower()):
                                break
                            city = str(input("select city name: "))
                            if not is_city(country.lower(), state.lower(), city.lower()):
                                continue
                            new_city = input_city()
                            data[country.lower()][state.lower()].remove(city)
                            if new_city in data[country.lower()][state.lower()]:
                                print('city is already exits please try again')
                                break
                            data[country.lower()][state.lower()].append(new_city)
                            print(data)
                            print("city updated......")
                            corn= continueOrNot()
                            if corn:
                                continue
                            else:
                                break
                    elif sub_choice == 4:
                        break
                    else:
                        print("Invalid choice. Please try again.")
                except ValueError:
                    print("Invalid input. Please enter a number.")
                    continue
        elif choice == 3:
            while True:
                print("""
                    ****Delete*****
                      Select option:
                        1. Country
                        2. State
                        3. City
                        4. Exit
                """)
                try:
                    sub_choice = int(input("Enter choice: "))
                    if sub_choice == 1:
                        while True:
                            Display_country()
                            if not is_country_empty(): 
                                break
                            country = str(input("select country name: "))
                            if not is_country(country.lower()):
                                continue
                            data.pop(country.lower())
                            print(data)
                            print("country deleted......")
                            corn= continueOrNot()
                            if corn:
                                continue
                            else:
                                break
                    elif sub_choice == 2:
                        while True:
                            Display_country()
                            if not is_country_empty(): 
                                break
                            country = str(input("select country name: "))
                            if not is_country(country.lower()):
                                continue
                            Display_state(country.lower())
                            if not is_state_empty(country.lower()):
                                break
                            state = str(input("select state name: "))
                            if not is_state(country.lower(), state.lower()):
                                continue
                            data[country.lower()].pop(state.lower())
                            print(data)
                            print("state deleted......")
                            corn= continueOrNot()
                            if corn:
                                continue
                            else:
                                break
                    elif sub_choice == 3:
                        while True:
                            Display_country()
                            if not is_country_empty(): 
                                break
                            country = str(input("select country name: "))
                            if not is_country(country.lower()):
                                continue
                            if not is_state_empty(country.lower()):
                                break
                            Display_state(country.lower())
                            state = str(input("select state name: "))
                            if not is_state(country.lower(), state.lower()):
                                continue    
                            if not is_city_empty(country.lower(), state.lower()):
                                break
                            Display_city(country.lower(), state.lower())
                            
                            city = str(input("select city name: "))
                            if not is_city(country.lower(), state.lower(), city.lower()):
                                continue
                            data[country.lower()][state.lower()].remove(city)
                            print(data)
                            print("city deleted......")
                            corn= continueOrNot()
                            if corn:
                                continue
                            else:
                                break
                    elif sub_choice == 4:
                        break
                    else:
                        print("Invalid choice. Please try again.")
                except ValueError:
                    print("Invalid input. Please enter a number.")
                    continue
        elif choice == 4:
            insert_data_into_excel(data, filename)
            break
        elif choice == 5:
            print(f"{data}")
        else:
            print("Invalid choice. Please try again.")
    except ValueError:
        print("Invalid input. Please enter a number.")
        continue
